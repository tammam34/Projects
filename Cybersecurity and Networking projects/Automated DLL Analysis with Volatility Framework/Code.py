import subprocess
import pandas as pd
import re
import os
import json
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER


class VolatilityDLLExtractor:

    def __init__(self, memory_image_path, volatility_path="vol", vol_version=3):
        self.memory_image = memory_image_path
        self.vol_path = volatility_path
        self.vol_version = vol_version
        self.output_dir = Path("volatility_output")
        self.output_dir.mkdir(exist_ok=True)

    def run_volatility_command(self, plugin, output_file=None):
        try:
            if self.vol_version == 3:
                cmd = [self.vol_path, "-f", self.memory_image, plugin]
            else:
                cmd = [self.vol_path, "-f", self.memory_image, "--profile=PROFILE", plugin]

            print(f"[*] Running: {' '.join(cmd)}")

            result = subprocess.run(cmd, capture_output=True, timeout=600)

            try:
                output = result.stdout.decode('utf-8', errors='replace')
                error = result.stderr.decode('utf-8', errors='replace')
            except Exception:
                try:
                    output = result.stdout.decode('latin-1', errors='replace')
                    error = result.stderr.decode('latin-1', errors='replace')
                except Exception:
                    output = str(result.stdout)
                    error = str(result.stderr)

            if result.returncode != 0 and error:
                print(f"[!] Volatility error: {error[:500]}")

            if output_file and output:
                output_path = self.output_dir / output_file
                with open(output_path, 'w', encoding='utf-8', errors='replace') as f:
                    f.write(output)
                print(f"[+] Output saved to: {output_path}")

            return output if output else ""

        except subprocess.TimeoutExpired:
            print(f"[!] Command timed out for plugin: {plugin}")
            return ""
        except FileNotFoundError:
            print(f"\n[!] ERROR: Volatility executable not found: '{self.vol_path}'")
            print("\n[!] Please ensure Volatility is installed. Installation options:")
            print("    ")
            print("    For Volatility 3:")
            print("      pip install volatility3")
            print("      Command: vol3 or python -m volatility3")
            print("    ")
            print("    For Volatility 2:")
            print("      Download from: https://www.volatilityfoundation.org/releases")
            print("      Command: python vol.py")
            print("    ")
            print("    Then run this script with:")
            print(f"      python {__file__} -i <memory_image> -v <volatility_path>")
            print("    ")
            print("    Examples:")
            print("      python test.py -i memory.dmp -v vol3")
            print("      python test.py -i memory.dmp -v 'python vol.py' --vol-version 2")
            print("      python test.py -i memory.dmp -v 'python -m volatility3'")
            return ""
        except Exception as e:
            print(f"[!] Error running {plugin}: {str(e)}")
            return ""

    def extract_dlllist(self):
        plugin = "windows.dlllist" if self.vol_version == 3 else "dlllist"
        return self.run_volatility_command(plugin, "dlllist_output.txt")


class DLLParser:

    def __init__(self):
        self.dll_data = []

    def parse_dlllist_vol3(self, output):
        lines = output.split('\n')
        current_process = None
        current_pid = None
        header_found = False

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if 'Volatility' in line or 'Framework' in line:
                continue

            if 'PID' in line and 'Process' in line and 'Base' in line:
                header_found = True
                continue

            if not header_found:
                continue

            parts = line.split()

            if len(parts) >= 2 and parts[0].isdigit():
                current_pid = parts[0]
                current_process = parts[1]

                if len(parts) >= 4 and '0x' in parts[2].lower():
                    base_addr = parts[2]
                    size = parts[3]

                    if len(parts) >= 5:
                        dll_name = parts[4]
                        dll_path = parts[5] if len(parts) > 5 else dll_name

                        for i, part in enumerate(parts):
                            if 'C:\\' in part or 'c:\\' in part.lower():
                                dll_path = part
                                if '\\' in dll_path:
                                    dll_name = dll_path.split('\\')[-1]
                                break

                        self.dll_data.append({
                            'PID': current_pid,
                            'Process': current_process,
                            'Base': base_addr,
                            'Size': size,
                            'Name': dll_name,
                            'Path': dll_path,
                            'LoadTime': 'N/A'
                        })

            elif current_process and '0x' in line.lower():
                parts = line.split()
                if len(parts) >= 3:
                    base_addr = parts[0]
                    size = parts[1]

                    dll_name = parts[2] if len(parts) > 2 else 'Unknown'
                    dll_path = dll_name

                    for part in parts:
                        if 'C:\\' in part or 'c:\\' in part.lower() or '\\' in part:
                            dll_path = part
                            if '\\' in dll_path:
                                dll_name = dll_path.split('\\')[-1]
                            break

                    if dll_path == dll_name and len(parts) > 3:
                        potential_path = ' '.join(parts[2:])
                        if 'C:\\' in potential_path or '\\' in potential_path:
                            dll_path = potential_path.split()[0]
                            if '\\' in dll_path:
                                dll_name = dll_path.split('\\')[-1]

                    self.dll_data.append({
                        'PID': current_pid,
                        'Process': current_process,
                        'Base': base_addr,
                        'Size': size,
                        'Name': dll_name,
                        'Path': dll_path,
                        'LoadTime': 'N/A'
                    })

        print(f"[*] Parsed {len(self.dll_data)} DLL entries")
        return self.dll_data

    def parse_dlllist_vol2(self, output):
        lines = output.split('\n')
        current_process = None
        current_pid = None

        for line in lines:
            if line.startswith('*' * 50):
                continue

            if 'Command line' in line or 'Process:' in line:
                match = re.search(r'pid:\s*(\d+)', line, re.IGNORECASE)
                if match:
                    current_pid = match.group(1)
                match = re.search(r'Process:\s*(\S+)', line, re.IGNORECASE)
                if match:
                    current_process = match.group(1)

            if '0x' in line and current_process:
                parts = re.split(r'\s{2,}', line.strip())
                if len(parts) >= 3:
                    self.dll_data.append({
                        'PID': current_pid,
                        'Process': current_process,
                        'Base': parts[0],
                        'Size': parts[1],
                        'Path': parts[2] if len(parts) > 2 else 'N/A',
                        'Name': os.path.basename(parts[2]) if len(parts) > 2 else 'Unknown',
                        'LoadTime': parts[3] if len(parts) > 3 else 'N/A'
                    })

        return self.dll_data

    def to_dataframe(self):
        return pd.DataFrame(self.dll_data)


class DLLAnalyzer:

    SUSPICIOUS_PATHS = [
        r'\\Temp\\',
        r'\\AppData\\Local\\Temp\\',
        r'\\Users\\[^\\]+\\Downloads\\',
        r'\\ProgramData\\',
        r'\\$Recycle\\.Bin\\',
        r'^C:\\Users\\Public\\',
        r'\\Content\\.Outlook\\',
    ]

    LEGITIMATE_PATHS = [
        r'C:\\Windows\\System32\\',
        r'C:\\Windows\\SysWOW64\\',
        r'C:\\Program Files\\',
        r'C:\\Program Files \(x86\)\\',
    ]

    CRITICAL_PROCESSES = [
        'lsass.exe', 'csrss.exe', 'wininit.exe', 'winlogon.exe',
        'services.exe', 'smss.exe', 'svchost.exe'
    ]

    def __init__(self, df):
        self.df = df
        self.findings = []

    def check_suspicious_paths(self):
        suspicious = []

        for idx, row in self.df.iterrows():
            path = row.get('Path', '')

            for pattern in self.SUSPICIOUS_PATHS:
                if re.search(pattern, path, re.IGNORECASE):
                    suspicious.append({
                        'PID': row['PID'],
                        'Process': row['Process'],
                        'DLL': row['Name'],
                        'Path': path,
                        'Reason': f'Suspicious path: {pattern}',
                        'Severity': 'HIGH'
                    })
                    break

        self.findings.extend(suspicious)
        return suspicious

    def check_critical_process_dlls(self):
        suspicious = []

        for idx, row in self.df.iterrows():
            process = row.get('Process', '').lower()
            path = row.get('Path', '')

            if any(proc in process for proc in self.CRITICAL_PROCESSES):
                is_legitimate = any(re.search(pattern, path, re.IGNORECASE)
                                  for pattern in self.LEGITIMATE_PATHS)

                if not is_legitimate and path and path != 'N/A':
                    suspicious.append({
                        'PID': row['PID'],
                        'Process': row['Process'],
                        'DLL': row['Name'],
                        'Path': path,
                        'Reason': f'Non-system DLL in critical process',
                        'Severity': 'CRITICAL'
                    })

        self.findings.extend(suspicious)
        return suspicious

    def check_unsigned_dlls(self):
        suspicious = []

        for idx, row in self.df.iterrows():
            path = row.get('Path', '')

            if path and path != 'N/A':
                is_system = any(re.search(pattern, path, re.IGNORECASE)
                              for pattern in self.LEGITIMATE_PATHS)

                if not is_system:
                    suspicious.append({
                        'PID': row['PID'],
                        'Process': row['Process'],
                        'DLL': row['Name'],
                        'Path': path,
                        'Reason': 'DLL not in standard system location',
                        'Severity': 'MEDIUM'
                    })

        self.findings.extend(suspicious)
        return suspicious

    def check_duplicate_system_dlls(self):
        suspicious = []

        system_dll_names = ['kernel32.dll', 'ntdll.dll', 'user32.dll',
                           'advapi32.dll', 'gdi32.dll', 'ws2_32.dll']

        for idx, row in self.df.iterrows():
            dll_name = row.get('Name', '').lower()
            path = row.get('Path', '')

            if dll_name in system_dll_names:
                is_system_path = any(re.search(pattern, path, re.IGNORECASE)
                                   for pattern in self.LEGITIMATE_PATHS)

                if not is_system_path and path and path != 'N/A':
                    suspicious.append({
                        'PID': row['PID'],
                        'Process': row['Process'],
                        'DLL': row['Name'],
                        'Path': path,
                        'Reason': f'System DLL "{dll_name}" loaded from non-system location',
                        'Severity': 'CRITICAL'
                    })

        self.findings.extend(suspicious)
        return suspicious

    def generate_statistics(self):
        stats = {
            'total_dlls': len(self.df),
            'unique_dlls': self.df['Name'].nunique(),
            'total_processes': self.df['PID'].nunique(),
            'unique_processes': self.df['Process'].nunique(),
            'suspicious_findings': len(self.findings),
            'critical_findings': sum(1 for f in self.findings if f.get('Severity') == 'CRITICAL'),
            'high_findings': sum(1 for f in self.findings if f.get('Severity') == 'HIGH'),
            'medium_findings': sum(1 for f in self.findings if f.get('Severity') == 'MEDIUM'),
        }

        stats['top_dlls'] = self.df['Name'].value_counts().head(10).to_dict()
        stats['top_processes'] = self.df['Process'].value_counts().head(10).to_dict()

        return stats

    def run_all_checks(self):
        print("[*] Checking suspicious paths...")
        self.check_suspicious_paths()

        print("[*] Checking critical processes...")
        self.check_critical_process_dlls()

        print("[*] Checking for unsigned/uncommon DLLs...")
        self.check_unsigned_dlls()

        print("[*] Checking for duplicate system DLLs...")
        self.check_duplicate_system_dlls()

        return self.findings


class ReportGenerator:

    def __init__(self, analysis_data):
        self.data = analysis_data
        self.df = pd.DataFrame(analysis_data['dll_dataframe'])
        self.findings = pd.DataFrame(analysis_data['findings']) if analysis_data['findings'] else pd.DataFrame()
        self.stats = analysis_data['statistics']

        self._clean_dataframes()

    def _clean_dataframes(self):
        import re

        def clean_string(val):
            if pd.isna(val):
                return ''
            if not isinstance(val, str):
                val = str(val)

            cleaned = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', val)

            cleaned = ''.join(char if ord(char) < 0xD800 or ord(char) > 0xDFFF else '?'
                            for char in cleaned)

            if len(cleaned) > 32767:
                cleaned = cleaned[:32767]

            return cleaned

        for col in self.df.columns:
            if self.df[col].dtype == 'object':
                self.df[col] = self.df[col].apply(clean_string)

        if not self.findings.empty:
            for col in self.findings.columns:
                if self.findings[col].dtype == 'object':
                    self.findings[col] = self.findings[col].apply(clean_string)

    def create_excel_report(self, output_file='dll_analysis_report.xlsx'):
        print(f"[*] Creating Excel report: {output_file}")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._create_summary_sheet(writer)
            self._create_findings_sheet(writer)
            self._create_all_dlls_sheet(writer)
            self._create_process_summary_sheet(writer)
            self._create_statistics_sheet(writer)

        print(f"[+] Excel report created: {output_file}")
        return output_file

    def _create_summary_sheet(self, writer):
        summary_data = [
            ['DLL Forensic Analysis Report', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Memory Image:', self.data.get('memory_image', 'N/A')],
            ['', ''],
            ['ANALYSIS SUMMARY', ''],
            ['Total DLLs Analyzed:', self.stats['total_dlls']],
            ['Unique DLLs:', self.stats['unique_dlls']],
            ['Total Processes:', self.stats['total_processes']],
            ['Unique Process Names:', self.stats['unique_processes']],
            ['', ''],
            ['FINDINGS', ''],
            ['Total Suspicious Findings:', self.stats['suspicious_findings']],
            ['Critical Severity:', self.stats['critical_findings']],
            ['High Severity:', self.stats['high_findings']],
            ['Medium Severity:', self.stats['medium_findings']],
        ]

        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        ws = writer.sheets['Summary']
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws['A1'].font = Font(size=16, bold=True)
        ws['A5'].font = Font(size=14, bold=True)
        ws['A11'].font = Font(size=14, bold=True)

    def _create_findings_sheet(self, writer):
        if not self.findings.empty:
            severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
            self.findings['severity_rank'] = self.findings['Severity'].map(severity_order)
            findings_sorted = self.findings.sort_values('severity_rank').drop('severity_rank', axis=1)

            findings_sorted.to_excel(writer, sheet_name='Suspicious Findings', index=False)

            ws = writer.sheets['Suspicious Findings']
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 12

            for row in range(2, len(findings_sorted) + 2):
                severity = ws[f'F{row}'].value
                if severity == 'CRITICAL':
                    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    ws[f'F{row}'].font = Font(color='FFFFFF', bold=True)
                elif severity == 'HIGH':
                    fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                    ws[f'F{row}'].font = Font(bold=True)
                elif severity == 'MEDIUM':
                    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                ws[f'F{row}'].fill = fill
        else:
            pd.DataFrame([['No suspicious findings detected']],
                        columns=['Status']).to_excel(writer, sheet_name='Suspicious Findings', index=False)

    def _create_all_dlls_sheet(self, writer):
        self.df.to_excel(writer, sheet_name='All DLLs', index=False)
        ws = writer.sheets['All DLLs']
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 60

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    def _create_process_summary_sheet(self, writer):
        process_summary = self.df.groupby(['PID', 'Process']).agg({
            'Name': 'count',
            'Path': 'nunique'
        }).reset_index()
        process_summary.columns = ['PID', 'Process', 'DLL Count', 'Unique Paths']
        process_summary = process_summary.sort_values('DLL Count', ascending=False)

        process_summary.to_excel(writer, sheet_name='Process Summary', index=False)

        ws = writer.sheets['Process Summary']
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_statistics_sheet(self, writer):
        stats_data = [
            ['Top 10 Most Common DLLs', ''],
            ['DLL Name', 'Count'],
        ]

        for dll, count in self.stats['top_dlls'].items():
            stats_data.append([dll, count])

        stats_data.extend([
            ['', ''],
            ['Top 10 Processes by DLL Count', ''],
            ['Process Name', 'DLL Count'],
        ])

        for proc, count in self.stats['top_processes'].items():
            stats_data.append([proc, count])

        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistics', index=False, header=False)

        ws = writer.sheets['Statistics']
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws['A1'].font = Font(size=14, bold=True)

    def create_pdf_report(self, output_file='dll_analysis_report.pdf'):
        print(f"[*] Creating PDF report: {output_file}")

        doc = SimpleDocTemplate(output_file, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("DLL Forensic Analysis Report", title_style))
        story.append(Spacer(1, 0.3*inch))

        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Paragraph(f"<b>Memory Image:</b> {self.data.get('memory_image', 'N/A')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        story.append(Paragraph("<b>Analysis Summary</b>", styles['Heading2']))
        stats_data = [
            ['Metric', 'Value'],
            ['Total DLLs Analyzed', str(self.stats['total_dlls'])],
            ['Unique DLLs', str(self.stats['unique_dlls'])],
            ['Total Processes', str(self.stats['total_processes'])],
            ['Suspicious Findings', str(self.stats['suspicious_findings'])],
            ['Critical Findings', str(self.stats['critical_findings'])],
            ['High Severity Findings', str(self.stats['high_findings'])],
            ['Medium Severity Findings', str(self.stats['medium_findings'])],
        ]

        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(stats_table)
        story.append(PageBreak())

        if not self.findings.empty:
            story.append(Paragraph("<b>Suspicious Findings</b>", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))

            findings_data = [['PID', 'Process', 'DLL', 'Severity', 'Reason']]

            for _, row in self.findings.head(20).iterrows():
                findings_data.append([
                    str(row['PID']),
                    str(row['Process'])[:20],
                    str(row['DLL'])[:25],
                    str(row['Severity']),
                    str(row['Reason'])[:40]
                ])

            findings_table = Table(findings_data, colWidths=[0.6*inch, 1.2*inch, 1.5*inch, 0.9*inch, 2.3*inch])
            findings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(findings_table)
        else:
            story.append(Paragraph("No suspicious findings detected.", styles['Normal']))

        doc.build(story)
        print(f"[+] PDF report created: {output_file}")
        return output_file


def main():
    parser = argparse.ArgumentParser(description='DLL Forensic Analysis Tool')
    parser.add_argument('-i', '--image', help='Path to memory image file')
    parser.add_argument('-v', '--volatility', default='vol', help='Path to Volatility executable')
    parser.add_argument('--vol-version', type=int, default=3, choices=[2, 3], help='Volatility version')
    parser.add_argument('-o', '--output-dir', default='.', help='Output directory for reports')
    parser.add_argument('--excel', action='store_true', help='Generate Excel report')
    parser.add_argument('--pdf', action='store_true', help='Generate PDF report')
    parser.add_argument('--all', action='store_true', help='Generate all reports')

    args = parser.parse_args()

    if not args.image:
        print("\n" + "=" * 70)
        print("DLL Forensic Analysis Tool - Interactive Mode")
        print("=" * 70)
        print("\nNo memory image specified. Please provide the required information:\n")

        args.image = input("Enter path to memory image file: ").strip().strip('"').strip("'")

        if not args.image:
            print("[!] Error: Memory image path is required.")
            return

        vol_path = input(f"Enter Volatility path (default: {args.volatility}): ").strip()
        if vol_path:
            args.volatility = vol_path

        vol_ver = input(f"Enter Volatility version - 2 or 3 (default: {args.vol_version}): ").strip()
        if vol_ver in ['2', '3']:
            args.vol_version = int(vol_ver)

        output_dir = input(f"Enter output directory (default: current directory): ").strip()
        if output_dir:
            args.output_dir = output_dir

        print("\n")

    if not args.excel and not args.pdf:
        args.all = True

    print("=" * 70)
    print("DLL Forensic Analysis Tool")
    print("=" * 70)

    if not os.path.exists(args.image):
        print(f"[!] Error: Memory image not found: {args.image}")
        return

    print("\n[+] Step 1: Extracting DLL data from memory image...")

    print("[*] Testing Volatility installation...")
    try:
        test_result = subprocess.run([args.volatility, '--help'],
                                    capture_output=True,
                                    timeout=10)
        if test_result.returncode != 0:
            print(f"[!] Warning: Volatility returned error code {test_result.returncode}")
        else:
            print("[+] Volatility found and working!")
    except FileNotFoundError:
        print(f"\n[!] ERROR: Cannot find Volatility at: '{args.volatility}'")
        print("\n[!] Common Volatility 3 commands to try:")
        print("    - vol3")
        print("    - vol")
        print("    - python -m volatility3")
        print("    - volatility3")
        print("\n[!] Please install Volatility 3:")
        print("    pip install volatility3")
        print("\n[!] Then run with the correct command:")
        print(f"    python test.py -i \"{args.image}\" -v vol3")
        return
    except Exception as e:
        print(f"[!] Error testing Volatility: {e}")

    extractor = VolatilityDLLExtractor(args.image, args.volatility, args.vol_version)
    dlllist_output = extractor.extract_dlllist()

    if not dlllist_output or len(dlllist_output) < 100:
        print("[!] Warning: Little or no output from Volatility.")
        return

    print("\n[+] Step 2: Parsing Volatility output...")
    parser_obj = DLLParser()

    if args.vol_version == 3:
        dll_data = parser_obj.parse_dlllist_vol3(dlllist_output)
    else:
        dll_data = parser_obj.parse_dlllist_vol2(dlllist_output)

    df = parser_obj.to_dataframe()

    if df.empty:
        print("[!] Error: No DLL data was parsed.")
        return

    print(f"[+] Parsed {len(df)} DLL entries from {df['PID'].nunique()} processes")

    print("\n[+] Step 3: Analyzing DLL data for suspicious indicators...")
    analyzer = DLLAnalyzer(df)
    findings = analyzer.run_all_checks()

    print(f"\n[+] Analysis complete. Found {len(findings)} suspicious indicators.")

    print("\n[+] Step 4: Generating statistics...")
    stats = analyzer.generate_statistics()

    print("\n" + "=" * 70)
    print("ANALYSIS SUMMARY")
    print("=" * 70)
    print(f"Total DLLs:           {stats['total_dlls']}")
    print(f"Unique DLLs:          {stats['unique_dlls']}")
    print(f"Total Processes:      {stats['total_processes']}")
    print(f"Suspicious Findings:  {stats['suspicious_findings']}")
    print(f"  - Critical:         {stats['critical_findings']}")
    print(f"  - High:             {stats['high_findings']}")
    print(f"  - Medium:           {stats['medium_findings']}")

    output_data = {
        'dll_dataframe': df.to_dict('records'),
        'findings': findings,
        'statistics': stats,
        'timestamp': datetime.now().isoformat(),
        'memory_image': args.image
    }

    json_file = os.path.join(args.output_dir, 'analysis_data.json')
    with open(json_file, 'w') as f:
        json.dump(output_data, f, indent=2)
    print(f"\n[+] Analysis data saved to {json_file}")

    report_gen = ReportGenerator(output_data)

    if args.excel or args.all:
        excel_file = os.path.join(args.output_dir, 'dll_analysis_report.xlsx')
        report_gen.create_excel_report(excel_file)

    if args.pdf or args.all:
        pdf_file = os.path.join(args.output_dir, 'dll_analysis_report.pdf')
        report_gen.create_pdf_report(pdf_file)

    print("\n[+] Analysis complete! Reports generated successfully.")


if __name__ == "__main__":
    main()